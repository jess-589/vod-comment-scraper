#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Step 1: Extract video links from a VOD tracking Excel workbook.

Expected sheet layout (one sheet per platform, e.g. TikTok / YouTube):
    Column B: account name
    Column D: video link
    Column E: publish date
    Column F: views
    Column G: comment count
    Column H: likes
    Column I: shares
    Data starts at row 3 (row 1-2 are headers).

If your workbook uses a different layout, adjust the constants below or
pass column letters via CLI options.
"""
import argparse
import json
import os

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def extract_sheet(ws, col_map, start_row):
    videos = []
    current_account = ""
    for row_idx in range(start_row, ws.max_row + 1):
        account = ws.cell(row=row_idx, column=col_map["account"]).value
        link = ws.cell(row=row_idx, column=col_map["link"]).value
        pub_date = ws.cell(row=row_idx, column=col_map["date"]).value
        views = ws.cell(row=row_idx, column=col_map["views"]).value
        comments = ws.cell(row=row_idx, column=col_map["comments"]).value
        likes = ws.cell(row=row_idx, column=col_map["likes"]).value
        shares = ws.cell(row=row_idx, column=col_map["shares"]).value

        if account:
            current_account = str(account)

        if link and str(link).strip():
            videos.append({
                "account": current_account,
                "link": str(link).strip(),
                "date": str(pub_date)[:10] if pub_date else "",
                "views": views or 0,
                "comment_count": comments or 0,
                "likes": likes or 0,
                "shares": shares or 0,
            })
    return videos


def run(excel_path, tt_sheet, yt_sheet, out_path,
        col_account="B", col_link="D", col_date="E",
        col_views="F", col_comments="G", col_likes="H", col_shares="I",
        start_row=3):
    col_map = {
        "account": column_index_from_string(col_account),
        "link": column_index_from_string(col_link),
        "date": column_index_from_string(col_date),
        "views": column_index_from_string(col_views),
        "comments": column_index_from_string(col_comments),
        "likes": column_index_from_string(col_likes),
        "shares": column_index_from_string(col_shares),
    }

    wb = load_workbook(excel_path, data_only=True)

    result = {}
    for platform, sheet_name in (("tiktok", tt_sheet), ("youtube", yt_sheet)):
        if not sheet_name or sheet_name not in wb.sheetnames:
            print(f"[extract] sheet '{sheet_name}' not found, skip {platform}")
            result[platform] = []
            continue
        result[platform] = extract_sheet(wb[sheet_name], col_map, start_row)
        vids = result[platform]
        with_comments = [v for v in vids
                         if isinstance(v["comment_count"], (int, float))
                         and v["comment_count"] > 0]
        print(f"[extract] {platform}: {len(vids)} videos "
              f"({len(with_comments)} have comments) from sheet '{sheet_name}'")

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[extract] saved -> {out_path}")
    return out_path


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Extract video links from VOD tracking Excel")
    p.add_argument("--excel", required=True, help="path to the .xlsx workbook")
    p.add_argument("--tt-sheet", default="数据追踪（TT）", help="TikTok sheet name")
    p.add_argument("--yt-sheet", default="数据追踪（YT）", help="YouTube sheet name")
    p.add_argument("--out", default=os.path.join("data", "video_links.json"))
    p.add_argument("--col-account", default="B")
    p.add_argument("--col-link", default="D")
    p.add_argument("--col-date", default="E")
    p.add_argument("--col-views", default="F")
    p.add_argument("--col-comments", default="G")
    p.add_argument("--col-likes", default="H")
    p.add_argument("--col-shares", default="I")
    p.add_argument("--start-row", type=int, default=3)
    args = p.parse_args()
    run(args.excel, args.tt_sheet, args.yt_sheet, args.out,
        args.col_account, args.col_link, args.col_date,
        args.col_views, args.col_comments, args.col_likes, args.col_shares,
        args.start_row)
