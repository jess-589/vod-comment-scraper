#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Step 3: Sentiment analysis + Excel report.

1. Loads YouTube + TikTok comment JSON files
2. VADER sentiment per comment (positive / negative / neutral)
3. Per-video aggregation: pos/neg/neu ratio, game keywords, key opinions
4. Writes a styled "评论原文及总结" sheet into an Excel workbook

If --base-excel is given, the original workbook is copied and the sheet is
appended to it; otherwise a brand-new workbook is created.
"""
import argparse
import json
import os
import shutil
from collections import Counter, defaultdict

# ============ Game keyword dictionary (edit for your game) ============
GAME_KEYWORDS = {
    # gameplay
    "gameplay": "玩法", "combat": "战斗", "pvp": "PVP", "pve": "PVE",
    "boss": "BOSS", "skill": "技能", "weapon": "武器", "character": "角色",
    "class": "职业", "build": "配装", "level": "等级", "difficulty": "难度",
    "mechanic": "机制", "loot": "掉落", "crafting": "制作", "skill tree": "技能树",
    "roguelike": "肉鸽", "dungeon": "地牢", "inventory": "背包", "gear": "装备",
    # audio-visual
    "graphic": "画面", "visual": "视觉", "art": "美术", "animation": "动画",
    "sound": "音效", "music": "音乐", "ui": "界面", "fps": "帧率", "lag": "卡顿",
    # sentiment expressions
    "fun": "好玩", "love": "喜欢", "hate": "讨厌", "boring": "无聊",
    "amazing": "惊艳", "cool": "酷", "awesome": "棒", "bad": "差",
    "good": "好", "great": "赞", "terrible": "糟糕", "broken": "崩了",
    "op": "超模", "nerf": "削弱", "buff": "加强", "meta": "环境",
    # IP / comparison
    "dark": "暗黑", "souls": "魂系", "elden": "老头环", "dnd": "DND",
    "medieval": "中世纪", "fantasy": "奇幻", "rpg": "RPG",
    # social
    "multiplayer": "多人", "co-op": "联机", "friend": "朋友", "team": "队伍",
    "streamer": "主播", "content": "内容", "update": "更新", "patch": "补丁",
}

POSITIVE_WORDS = {
    "love", "amazing", "awesome", "great", "fun", "cool", "best", "incredible",
    "perfect", "excited", "hype", "insane", "fire", "goated", "masterpiece",
    "beautiful", "gorgeous", "stunning", "addicted", "hooked", "can't wait",
    "underrated", "slept", "banger", "w", "wlr", "based", "chad",
}

NEGATIVE_WORDS = {
    "hate", "trash", "garbage", "boring", "bad", "terrible", "worst", "awful",
    "disappointed", "broken", "buggy", "dead", "scam", "ripoff", "lame",
    "stupid", "dumb", "mid", "cooked", "ass", "shit", "fuck", "sucks",
    "unplayable", "crash", "lag", "pay to win", "p2w", "flop",
}

_analyzer = None


def analyze_sentiment(text):
    """VADER sentiment for a single comment."""
    global _analyzer
    if _analyzer is None:
        from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
        _analyzer = SentimentIntensityAnalyzer()

    text_lower = text.lower()
    boost = sum(1 for w in POSITIVE_WORDS if w in text_lower)
    penalty = sum(1 for w in NEGATIVE_WORDS if w in text_lower)
    compound = _analyzer.polarity_scores(text)["compound"]
    compound = max(-1.0, min(1.0, compound + 0.05 * boost - 0.05 * penalty))

    if compound >= 0.05:
        label = "正面"
    elif compound <= -0.05:
        label = "负面"
    else:
        label = "中性"
    return {"label": label, "compound": round(compound, 3)}


def extract_keywords(text):
    text_lower = text.lower()
    return [f"{en}({zh})" for en, zh in GAME_KEYWORDS.items() if en in text_lower]


def extract_key_opinions(comments, label, top=5):
    opinions = []
    for c in comments:
        if c["sentiment"]["label"] == label:
            t = c["comment"]
            if len(t) > 100:
                t = t[:100] + "..."
            opinions.append(t)
    return opinions[:top]


def group_by_video(comments):
    groups = defaultdict(lambda: {"comments": []})
    for c in comments:
        g = groups[(c["platform"], c["video_url"])]
        g["platform"] = c["platform"]
        g["account"] = c.get("account", "")
        g["video_title"] = c.get("video_title", "")
        g["video_url"] = c["video_url"]
        g["comments"].append(c)
    return groups


def write_to_excel(comments, video_groups, output_path, base_excel, title, sheet_name):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if base_excel and os.path.exists(base_excel):
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        shutil.copy2(base_excel, output_path)
        wb = load_workbook(output_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=14, color="FFFFFF", name="Microsoft YaHei")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="FFFFFF", name="Microsoft YaHei")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Microsoft YaHei")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    neu_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin", color="D9D9D9"),
                  right=Side(style="thin", color="D9D9D9"),
                  top=Side(style="thin", color="D9D9D9"),
                  bottom=Side(style="thin", color="D9D9D9"))

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f"{title} 评论原文及情感分析")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = header_align
    ws.row_dimensions[row].height = 30
    row += 2

    # ---- Section 1: raw comments ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value="一、评论原文")
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = header_align
    row += 1

    headers = ["平台", "账号", "视频标题", "视频链接", "评论原文", "情感倾向", "情感分值"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin
    row += 1

    sorted_comments = sorted(
        comments,
        key=lambda x: (0 if x["platform"] == "TikTok" else 1,
                       x.get("account", ""), x["video_url"]))
    for c in sorted_comments:
        ws.cell(row=row, column=1, value=c["platform"])
        ws.cell(row=row, column=2, value=c.get("account", ""))
        ws.cell(row=row, column=3, value=c.get("video_title", ""))
        ws.cell(row=row, column=4, value=c["video_url"])
        ws.cell(row=row, column=5, value=c["comment"])
        label = c["sentiment"]["label"]
        cf = ws.cell(row=row, column=6, value=label)
        cf.fill = pos_fill if label == "正面" else neg_fill if label == "负面" else neu_fill
        cg = ws.cell(row=row, column=7, value=c["sentiment"]["compound"])
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = (Alignment(horizontal="center", vertical="center", wrap_text=True)
                              if col >= 6 else cell_align)
        row += 1

    # ---- Section 2: per-video analysis ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value="二、玩家偏好及正负向态度分析（按视频汇总）")
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = header_align
    row += 1

    analysis_headers = ["平台", "账号", "视频标题", "视频链接", "评论数",
                        "正面占比", "负面占比", "中性占比",
                        "玩家偏好关键词", "关键正面观点", "关键负面观点"]
    for col, h in enumerate(analysis_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin
    row += 1

    for (platform, video_url), info in sorted(
            video_groups.items(),
            key=lambda x: (0 if x[1]["platform"] == "TikTok" else 1,
                           x[1].get("account", ""))):
        cs = info["comments"]
        total = len(cs)
        if total == 0:
            continue
        pos = sum(1 for c in cs if c["sentiment"]["label"] == "正面")
        neg = sum(1 for c in cs if c["sentiment"]["label"] == "负面")
        neu = total - pos - neg
        kws = Counter(kw for c in cs for kw in c.get("keywords", []))
        pos_ops = " | ".join(extract_key_opinions(cs, "正面", 3)) or "无"
        neg_ops = " | ".join(extract_key_opinions(cs, "负面", 3)) or "无"

        ws.cell(row=row, column=1, value=info["platform"])
        ws.cell(row=row, column=2, value=info.get("account", ""))
        ws.cell(row=row, column=3, value=info.get("video_title", ""))
        ws.cell(row=row, column=4, value=info["video_url"])
        ws.cell(row=row, column=5, value=total)
        ws.cell(row=row, column=6, value=f"{pos}条 ({pos / total * 100:.0f}%)")
        ws.cell(row=row, column=7, value=f"{neg}条 ({neg / total * 100:.0f}%)")
        ws.cell(row=row, column=8, value=f"{neu}条 ({neu / total * 100:.0f}%)")
        ws.cell(row=row, column=9, value="、".join(k for k, _ in kws.most_common(10)) or "无")
        ws.cell(row=row, column=10, value=pos_ops)
        ws.cell(row=row, column=11, value=neg_ops)
        if pos > neg:
            ws.cell(row=row, column=6).fill = pos_fill
        if neg > pos:
            ws.cell(row=row, column=7).fill = neg_fill
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin
        row += 1

    # ---- Section 3: overall summary ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value="三、整体总结")
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = header_align
    row += 1

    total = len(comments)
    pos = sum(1 for c in comments if c["sentiment"]["label"] == "正面")
    neg = sum(1 for c in comments if c["sentiment"]["label"] == "负面")
    neu = total - pos - neg
    videos_n = len([g for g in video_groups.values() if g["comments"]])

    stats = [
        ("评论总条数", str(total)),
        ("有评论的视频数", str(videos_n)),
        ("正面评论", f"{pos}条 ({pos / total * 100:.1f}%)" if total else "0条"),
        ("负面评论", f"{neg}条 ({neg / total * 100:.1f}%)" if total else "0条"),
        ("中性评论", f"{neu}条 ({neu / total * 100:.1f}%)" if total else "0条"),
    ]
    for label, value in stats:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True, name="Microsoft YaHei", size=10)
        c1.alignment = cell_align
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = cell_font
        c2.alignment = cell_align
        row += 1

    row += 1
    all_kws = Counter(kw for c in comments for kw in c.get("keywords", []))
    c1 = ws.cell(row=row, column=1, value="玩家高频关注词")
    c1.font = Font(bold=True, name="Microsoft YaHei", size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c2 = ws.cell(row=row, column=2,
                 value="、".join(f"{k}({n}次)" for k, n in all_kws.most_common(30)) or "无")
    c2.font = cell_font
    c2.alignment = cell_align
    row += 2

    all_pos = extract_key_opinions(comments, "正面", 10)
    all_neg = extract_key_opinions(comments, "负面", 10)
    c1 = ws.cell(row=row, column=1, value="代表性正面观点")
    c1.font = Font(bold=True, name="Microsoft YaHei", size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c2 = ws.cell(row=row, column=2, value=" | ".join(all_pos) or "无")
    c2.font = cell_font
    c2.alignment = cell_align
    c2.fill = pos_fill
    row += 1

    c1 = ws.cell(row=row, column=1, value="代表性负面观点")
    c1.font = Font(bold=True, name="Microsoft YaHei", size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c2 = ws.cell(row=row, column=2, value=" | ".join(all_neg) or "无")
    c2.font = cell_font
    c2.alignment = cell_align
    c2.fill = neg_fill

    for col, width in {"A": 10, "B": 18, "C": 35, "D": 45, "E": 60, "F": 10,
                       "G": 10, "H": 12, "I": 12, "J": 12, "K": 40,
                       "L": 50, "M": 50}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"[analyze] Excel saved -> {output_path}")


def run(yt_path, tt_path, output_path, base_excel, title, sheet_name):
    comments = []
    for name, path in (("YouTube", yt_path), ("TikTok", tt_path)):
        if path and os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            comments.extend(data)
            print(f"[analyze] {name}: {len(data)} comments")
        else:
            print(f"[analyze] {name}: file not found ({path}), skipped")

    if not comments:
        raise SystemExit("[analyze] ERROR: no comment data found. "
                         "Run the scraping steps first.")

    for i, c in enumerate(comments):
        text = (c.get("comment") or "").strip()
        if not text:
            c["sentiment"] = {"label": "中性", "compound": 0}
            c["keywords"] = []
            continue
        c["sentiment"] = analyze_sentiment(text)
        c["keywords"] = extract_keywords(text)
        if (i + 1) % 200 == 0:
            print(f"[analyze] {i + 1}/{len(comments)} analyzed ...")

    video_groups = group_by_video(comments)
    print(f"[analyze] videos with comments: {len(video_groups)}")
    write_to_excel(comments, video_groups, output_path, base_excel, title, sheet_name)
    return output_path


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Sentiment analysis + Excel report")
    p.add_argument("--yt-comments", default=os.path.join("data", "youtube_comments.json"))
    p.add_argument("--tt-comments", default=os.path.join("data", "tiktok_comments.json"))
    p.add_argument("--output", required=True, help="output .xlsx path")
    p.add_argument("--base-excel", default=None,
                   help="optional: copy this workbook and append the sheet to it")
    p.add_argument("--title", default="Game", help="game name used in report title")
    p.add_argument("--sheet", default="评论原文及总结")
    args = p.parse_args()
    run(args.yt_comments, args.tt_comments, args.output,
        args.base_excel, args.title, args.sheet)
